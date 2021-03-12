import pyautogui as pt
from time import sleep
import pyperclip
import random
import openpyxl as op
from datetime import datetime
import webbrowser

webbrowser.open("https://web.whatsapp.com/")
now = datetime.now()
sleep(20)
flag = 0
try:
    pos1 = pt.locateOnScreen("whatsapp/EmojiAttachment.PNG", confidence=.6)
    x = pos1[0]
    y = pos1[1]
except Exception:
    pos1 = pt.locateOnScreen("whatsapp/home.PNG", confidence=.5)
    x = pos1[0]
    y = pos1[1]


def get_message(name):
    # print("Get message running...")
    sleep(2)
    global x, y
    if flag == 0:
        pos = pt.locateOnScreen("whatsapp/EmojiAttachment.PNG", confidence=.6)
        x = pos[0]
        y = pos[1]
        pt.moveTo(x, y, duration=.05)
        pt.moveTo(x + 120, y - 60, duration=.5)
        pt.tripleClick()
        pt.rightClick()
        pt.moveRel(40, -170)
        pt.click()
        msg = pyperclip.paste()
        pt.click()
        print("\t\t\t\t", name, "Send Message as: ", msg)
    else:
        print("\t\t\t\t", name, "Group Send Message")
    # print("Get message running...")
    return name


def post_response(message):
    global x, y
    pos = pt.locateOnScreen("whatsapp/EmojiAttachment.PNG", confidence=.6)
    x = pos[0]
    y = pos[1]
    pt.moveTo(x + 150, y + 20, duration=.5)
    pt.click()
    pt.typewrite(message, interval=.01)
    pt.typewrite("\n", interval=.01)
    # print("Post response running...")


def take_name():
    global x, y, flag
    pos = pt.locateOnScreen("whatsapp/search.PNG", confidence=.9)
    x = pos[0]
    y = pos[1]
    pt.moveTo(x - 500, y + 5, duration=.5)
    pt.click()
    sleep(2)
    try:
        pos2 = pt.locateOnScreen("whatsapp/Contact.PNG", confidence=.9)
        x = pos2[0]
        y = pos2[1]
        pt.moveTo(x - 40, y + 380, duration=.5)
    except Exception:
        pos2 = pt.locateOnScreen("whatsapp/Group.PNG", confidence=.8)
        flag = 1
        # print(flag)
        x = pos2[0]
        y = pos2[1]
        pt.moveTo(x - 35, y + 390, duration=.5)
    pt.tripleClick()
    pt.rightClick()
    pt.moveRel(0, 15)
    pt.click()
    name = pyperclip.paste()
    pt.moveTo(x - 40, y + 10)
    pt.click()
    # print("Take name running...")
    return name


# def process_response(message, name):
def process_response(name):
    random_no = random.randrange(3)
    # print("Process Response running...")
    # if "?" in str(message).lower():
    if flag == 1:
        if random_no == 0:
            return "Hey {} group members, Ayan is not here right now, he will contact ASAP!!".format(name)
        elif random_no == 1:
            return "Greeting of the day to all {} group members!! Ayan will soon respond!!".format(name)
        else:
            return "Dear {} group members, Ayan will revert back ASAP!!".format(name)
    else:
        # if name == 'Dadi Wala Baby':
        #     return "Hey {}, I'm bipeen Lord, Ayan's bot!! Ayan is busy somewhere. He will text you asap".format(name)
        if name == 'Akash':
            return "Hello {}, I'm Akash and I'm a Bot!! Ayan is currently not here, soon he will contact".format(name)
        else:
            return "Greetings to you {}!! I'm a Bot! Currently I'm receiving Ayan's message!".format(name)


def check_message():
    # pt.moveTo(x + 50, y - 43, duration=.5)
    global flag
    while True:
        try:
            pos = pt.locateOnScreen("whatsapp/green.PNG", confidence=.7)
            if pos is not None:
                pt.moveTo(pos)
                pt.moveRel(-100, 0)
                pt.click()
                sleep(.5)
                flag = 0
                # msg, name = get_message(take_name())
                # response = process_response(msg, name)
                name = get_message(take_name())
                response = process_response(name)
                post_response(response)
                insert(name)
            else:
                print("No new message yet....")
        except Exception:
            print("No new message!!")
        # print("Check Message running...")
        sleep(5)
        # if pt.pixelMatchesColor(int(x + 50), int(y - 35), (38, 45, 49), tolerance=10):
        #     print("Is Grey!!")
        # else:
        #     print("No new message yet....")


def insert(name):
    path = 'C:\\Users\\AYAN\\PycharmProjects\\WhatsappBot\\data.xlsx'
    wb = op.load_workbook(path)
    sheet = wb.active
    rw = sheet.max_row
    sheet.cell(rw + 1, 1).value = name
    sheet.cell(rw + 1, 2).value = now.strftime("%d-%m-%Y  %H:%M:%S")
    wb.save('data.xlsx')
    # i = 1
    # while True:
    #     if sheet.cell(i, 1).value is None:
    #         print("running if insert...", sheet.cell(i, 1).value)
    #         sheet.cell(i, 1).value = name
    #         print(sheet.cell(i, 1).value)
    #         wb.save("data.xlsx")
    #         break
    #     else:
    #         print("running else insert...", sheet.cell(i, 1).value)
    #         i = i+1


check_message()
